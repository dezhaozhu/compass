from typing import List
from datetime import timedelta
import pandas as pd

# from camel.logger import get_logger
# from camel.toolkits.base import BaseToolkit
# from camel.toolkits.function_tool import FunctionTool

# logger = get_logger(__name__)


class ExcelToolkit():
# class ExcelToolkit(BaseToolkit):
    r"""A class representing a toolkit for extract detailed cell information
    from an Excel file.

    This class provides method for processing docx, pdf, pptx, etc. It cannot
    process excel files.
    """

    def _convert_to_markdown(self, df: pd.DataFrame) -> str:
        r"""Convert DataFrame to Markdown format table.

        Args:
            df (pd.DataFrame): DataFrame containing the Excel data.

        Returns:
            str: Markdown formatted table.
        """
        from tabulate import tabulate

        md_table = tabulate(df, headers='keys', tablefmt='pipe')
        return str(md_table)

    def _process_sheet(self,human_preference,manufacturing_ability,df):
        df = df.sort_values(by=['计划完工时间', 'WBS元素'], ascending=[True, True])

        df['AI推算分包单位'] = ''
        df['AI推算完工时间'] = None
        df['AI推算完工情况'] = ''

        df['计划完工时间'] = pd.to_datetime(df['计划完工时间'])
        df['计划完工时间'] = df['计划完工时间'] - timedelta(days=7)

        for index, row in df.iterrows():
            wbs = row['WBS元素']
            if wbs in human_preference:
                df.at[index, 'AI推算分包单位'] = human_preference[wbs]

        for index, row in df.iterrows():
                subcontractor = row['AI推算分包单位']
                if subcontractor and subcontractor in manufacturing_ability:
                    first_key = next(iter(manufacturing_ability[subcontractor])) # 简化了任务

                    line_bend_ability = manufacturing_ability[subcontractor][first_key]['线弯']
                    single_bend_ability = manufacturing_ability[subcontractor][first_key]['单机弯']
                    
                    if pd.notna(row['实际投料日期']):
                        line_bend = row.get('线弯', 0)
                        single_bend = row.get('单机弯', 0)
                        try:
                            line_bend = float(line_bend) if line_bend else 0
                            single_bend = float(single_bend) if single_bend else 0
                            
                            extra_days = 0
                            if line_bend_ability > 0:
                                extra_days += line_bend / line_bend_ability
                            if single_bend_ability > 0:
                                extra_days += single_bend / single_bend_ability
                            
                            completion_date = row['实际投料日期'] + timedelta(days=extra_days)
                            df.at[index, 'AI推算完工时间'] = completion_date
                            
                            if completion_date <= row['计划完工时间']:
                                df.at[index, 'AI推算完工情况'] = '提前完工'
                            else:
                                df.at[index, 'AI推算完工情况'] = '滞后完工'
                        except (ValueError, TypeError):
                            continue
        return df

    def extract_excel_content(self, human_preference, manufacturing_ability, document_path: str) -> str:
        r"""Extract detailed cell information from an Excel file, including
        multiple sheets.

        Args:
            document_path (str): The path of the Excel file.

        Returns:
            str: Extracted excel information, including details of each sheet.
        """
        from openpyxl import load_workbook
        from xls2xlsx import XLS2XLSX

        # logger.debug(
        #     f"Calling extract_excel_content with document_path"
        #     f": {document_path}"
        # )

        if not (
            document_path.endswith("xls")
            or document_path.endswith("xlsx")
            or document_path.endswith("csv")
        ):
            # logger.error("Only xls, xlsx, csv files are supported.")
            return (
                f"Failed to process file {document_path}: "
                f"It is not excel format. Please try other ways."
            )

        if document_path.endswith("csv"):
            try:
                df = pd.read_csv(document_path)
                # md_table = self._convert_to_markdown(df)
                # return f"CSV File Processed:\n{md_table}"
                # return f"CSV File has been processed."
            except Exception as e:
                # logger.error(f"Failed to process file {document_path}: {e}")
                return f"Failed to process file {document_path}: {e}"

        if document_path.endswith("xls"):
            output_path = document_path.replace(".xls", ".xlsx")
            x2x = XLS2XLSX(document_path)
            x2x.to_xlsx(output_path)
            document_path = output_path

        # Load the Excel workbook
        wb = load_workbook(document_path, data_only=True)
        sheet_info_list = []

        # Iterate through all sheets
        for sheet in wb.sheetnames:
            # ws = wb[sheet]
            # cell_info_list = []

            # for row in ws.iter_rows():
            #     for cell in row:
            #         row_num = cell.row
            #         col_letter = cell.column_letter

            #         cell_value = cell.value

            #         font_color = None
            #         if (
            #             cell.font
            #             and cell.font.color
            #             and "rgb=None" not in str(cell.font.color)
            #         ):  # Handle font color
            #             font_color = cell.font.color.rgb

            #         fill_color = None
            #         if (
            #             cell.fill
            #             and cell.fill.fgColor
            #             and "rgb=None" not in str(cell.fill.fgColor)
            #         ):  # Handle fill color
            #             fill_color = cell.fill.fgColor.rgb

            #         cell_info_list.append(
            #             {
            #                 "index": f"{row_num}{col_letter}",
            #                 "value": cell_value,
            #                 "font_color": font_color,
            #                 "fill_color": fill_color,
            #             }
            #         )

            # Convert the sheet to a DataFrame and then to markdown
            sheet_df = pd.read_excel(document_path, sheet_name=sheet, engine='openpyxl')
            if all(item in sheet_df.columns for item in ["WBS元素","实际投料日期","计划完工时间","备注"]):
                df = self._process_sheet(human_preference,manufacturing_ability,sheet_df)
            else:
                continue
            # markdown_content = self._convert_to_markdown(sheet_df)

            # Collect all information for the sheet
            # sheet_info = {
            #     "sheet_name": sheet,
            #     "cell_info_list": cell_info_list,
            #     "markdown_content": markdown_content,
            # }
            # sheet_info_list.append(sheet_info)

        # result_str = ""
        # for sheet_info in sheet_info_list:
        #     result_str += f"""
        #     Sheet Name: {sheet_info['sheet_name']}
        #     Cell information list:
        #     {sheet_info['cell_info_list']}
            
        #     Markdown View of the content:
        #     {sheet_info['markdown_content']}
            
        #     {'-'*40}
        #     """

        return df


    # def get_tools(self) -> List[FunctionTool]:
    #     r"""Returns a list of FunctionTool objects representing the functions
    #     in the toolkit.

    #     Returns:
    #         List[FunctionTool]: A list of FunctionTool objects representing
    #             the functions in the toolkit.
    #     """
    #     return [
    #         FunctionTool(self.extract_excel_content),
    #     ]